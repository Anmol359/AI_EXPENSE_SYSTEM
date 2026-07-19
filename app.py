from flask import Flask, render_template, request, redirect, session
import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
import os
from flask import send_file
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl import Workbook
from flask import send_file
import traceback

app = Flask(__name__)

app.secret_key = "AI_EXPENSE_SYSTEM_2026"

import os

conn = sqlite3.connect("finance_system.db", check_same_thread=False)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()


# ---------------- HOME ----------------

@app.route("/")
def home():
    return render_template("home.html")


# ---------------- REGISTER ----------------

@app.route("/register", methods=["GET", "POST"])
def register():

    if request.method == "POST":

        name = request.form["name"]
        email = request.form["email"]
        password = request.form["password"]

        sql = """
        INSERT INTO users(name,email,password)
        VALUES(?,?,?)
        """

        values = (name, email, password)

        cursor.execute(sql, values)
        conn.commit()

        return redirect("/login")

    return render_template("register.html")


# ---------------- LOGIN ----------------

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        email = request.form["email"]
        password = request.form["password"]

        sql = """
        SELECT * FROM users
        WHERE email=? AND password=?
        """

        values = (email, password)

        cursor.execute(sql, values)

        user = cursor.fetchone()

        if user:

            session["user_id"] = user[0]
            session["user_name"] = user[1]

            return redirect("/dashboard")

        else:

           return render_template(
            "login.html",
            error="Invalid Email or Password"
)

    return render_template("login.html")


# ---------------- LOGOUT ----------------

@app.route("/logout")
def logout():

    session.clear()

    return redirect("/login")

# ---------------- DASHBOARD ----------------

@app.route("/dashboard")
def dashboard():

    if "user_id" not in session:
        return redirect("/login")

    mycursor = conn.cursor()

    # Total Expense
    sql = """
    SELECT COALESCE(SUM(amount),0)
    FROM expenses
    WHERE user_id=?
    """

    mycursor.execute(sql, (session["user_id"],))

    total = float(mycursor.fetchone()[0])

    # Total Transactions
    sql = """
    SELECT COUNT(*)
    FROM expenses
    WHERE user_id=?
    """

    mycursor.execute(sql, (session["user_id"],))

    transaction_count = mycursor.fetchone()[0]


# Budget

    sql = """
    SELECT monthly_budget
    FROM budget
    WHERE user_id=?
    ORDER BY budget_id DESC
    LIMIT 1
    """

    mycursor.execute(sql, (session["user_id"],))

    result = mycursor.fetchone()

    if result:
        budget = float(result[0])
    else:
        budget = 0

    remaining = budget - total
# ---------------- AI Financial Advisor ----------------

    if budget == 0:

     ai_message = "⚠️ Please set your monthly budget first."

    elif total <= budget * 0.50:

     ai_message = (
        "✅ Excellent! You are spending wisely. "
        "Keep saving at this pace."
    )

    elif total <= budget * 0.80:

     ai_message = (
        "🙂 Your spending is under control. "
        "Avoid unnecessary shopping to save more."
    )

    elif total <= budget:

     ai_message = (
        "⚠️ You have used more than 80% of your budget. "
        "Reduce Food and Shopping expenses for the rest of the month."
    )

    else:

     ai_message = (
        "🚨 Budget Exceeded! "
        "Next month try limiting Food, Shopping and Entertainment expenses."
    )
    mycursor.close()

    return render_template(
    "dashboard.html",
    username=session["user_name"],
    total=total,
    budget=budget,
    remaining=remaining,
    transaction_count=transaction_count,
    ai_message=ai_message
)
# ---------------- ADD EXPENSE ----------------

@app.route("/add_expense", methods=["GET", "POST"])
def add_expense():

    if "user_id" not in session:
        return redirect("/login")

    if request.method == "POST":

        amount = request.form["amount"]
        category = request.form["category"]
        description = request.form["description"]
        expense_date = request.form["expense_date"]

        mycursor = conn.cursor()

        sql = """
        INSERT INTO expenses
        (user_id,amount,category,description,expense_date)
        VALUES(?,?,?,?,?)
        """

        values = (
            session["user_id"],
            amount,
            category,
            description,
            expense_date
        )

        mycursor.execute(sql, values)
        conn.commit()

        mycursor.close()

        return redirect("/expenses")

    return render_template("add_expense.html")


# ---------------- EXPENSE HISTORY ----------------

@app.route("/expenses")
def expenses():

    if "user_id" not in session:
        return redirect("/login")

    mycursor = conn.cursor()

    search = request.args.get("search")

    if search:

        sql = """
        SELECT *
        FROM expenses
        WHERE user_id=?
        AND (category LIKE ? OR description LIKE ?)
        ORDER BY expense_id DESC
        """

        value = "%" + search + "%"

        mycursor.execute(
            sql,
            (
                session["user_id"],
                value,
                value
            )
        )

    else:

        sql = """
        SELECT *
        FROM expenses
        WHERE user_id=?
        ORDER BY expense_id DESC
        """

        mycursor.execute(
            sql,
            (
                session["user_id"],
            )
        )

    data = mycursor.fetchall()

    mycursor.close()

    return render_template(
        "expense_history.html",
        expenses=data
    )


# ---------------- EDIT EXPENSE ----------------

@app.route("/edit_expense/<int:id>", methods=["GET", "POST"])
def edit_expense(id):

    if "user_id" not in session:
        return redirect("/login")

    mycursor = conn.cursor()

    if request.method == "POST":

        amount = request.form["amount"]
        category = request.form["category"]
        description = request.form["description"]
        expense_date = request.form["expense_date"]

        sql = """
        UPDATE expenses
        SET amount=?,
            category=?,
            description=?,
            expense_date=?
        WHERE expense_id=?
        AND user_id=?
        """

        values = (
            amount,
            category,
            description,
            expense_date,
            id,
            session["user_id"]
        )

        mycursor.execute(sql, values)
        conn.commit()
        mycursor.close()

        return redirect("/expenses")

    sql = """
    SELECT *
    FROM expenses
    WHERE expense_id=?
    AND user_id=?
    """

    mycursor.execute(sql, (id, session["user_id"]))

    expense = mycursor.fetchone()

    mycursor.close()

    return render_template(
        "edit_expense.html",
        expense=expense
    )


# ---------------- DELETE EXPENSE ----------------

@app.route("/delete_expense/<int:id>")
def delete_expense(id):

    if "user_id" not in session:
        return redirect("/login")

    mycursor = conn.cursor()

    sql = """
    DELETE FROM expenses
    WHERE expense_id=?
    AND user_id=?
    """

    mycursor.execute(sql, (id, session["user_id"]))

    conn.commit()

    mycursor.close()

    return redirect("/expenses")    

# ---------------- BUDGET ----------------

@app.route("/budget", methods=["GET", "POST"])
def budget():

    if "user_id" not in session:
        return redirect("/login")

    mycursor = conn.cursor()

    if request.method == "POST":

        budget = request.form["budget"]

        sql = """
        INSERT INTO budget(user_id,monthly_budget,month,year)
        VALUES(?,?,?,?)
        """

        from datetime import datetime

        today = datetime.now()

        mycursor.execute(
          sql,
          (
            session["user_id"],
            budget,
            today.month,
            today.year
          )
        )

        conn.commit()

        mycursor.close()

        return redirect("/dashboard")

    mycursor.close()

    return render_template("budget.html")



@app.route("/analytics")
def analytics():

    if "user_id" not in session:
        return redirect("/login")

    chart = request.args.get("chart", "pie")

    # Category Data
    query = """
    SELECT category, SUM(amount) AS total
    FROM expenses
    WHERE user_id=?
    GROUP BY category
    """

    df = pd.read_sql(
        query,
        conn,
        params=(session["user_id"],)
    )

    # Daily Data
    line_query = """
    SELECT expense_date, SUM(amount) AS total
    FROM expenses
    WHERE user_id=?
    GROUP BY expense_date
    ORDER BY expense_date
    """

    df_line = pd.read_sql(
        line_query,
        conn,
        params=(session["user_id"],)
    )

    # Generate Selected Chart
    if chart == "pie":

        plt.figure(figsize=(6,6))
        plt.pie(
            df["total"],
            labels=df["category"],
            autopct="%1.1f%%",
            startangle=90
        )
        plt.title("Expense By Category")

    elif chart == "bar":

        plt.figure(figsize=(8,5))
        plt.bar(df["category"], df["total"])
        plt.title("Category Wise Expense")
        plt.xlabel("Category")
        plt.ylabel("Amount")

    elif chart == "line":

        plt.figure(figsize=(8,5))
        plt.plot(
            df_line["expense_date"],
            df_line["total"],
            marker="o",
            linewidth=3
        )
        plt.title("Daily Expense Trend")
        plt.xlabel("Date")
        plt.ylabel("Amount")
        plt.xticks(rotation=45)

    elif chart == "donut":

        plt.figure(figsize=(6,6))
        plt.pie(
            df["total"],
            labels=df["category"],
            autopct="%1.1f%%",
            startangle=90,
            wedgeprops=dict(width=0.45)
        )
        plt.title("Expense Donut Chart")

    plt.tight_layout()
    plt.savefig("static/charts/chart.png")
    plt.close()

    return render_template(
        "analytics.html",
        chart=chart
    )


@app.route("/export_pdf")
def download_pdf():

    if "user_id" not in session:
        return redirect("/login")

    mycursor = conn.cursor()

    # Total Expense
    mycursor.execute(
        "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE user_id=?",
        (session["user_id"],)
    )
    total = float(mycursor.fetchone()[0])

    # Budget
    mycursor.execute(
        """
        SELECT monthly_budget
        FROM budget
        WHERE user_id=?
        ORDER BY budget_id DESC
        LIMIT 1
        """,
        (session["user_id"],)
    )

    result = mycursor.fetchone()
    budget = float(result[0]) if result else 0

    remaining = budget - total

    pdf_file = "expense_report.pdf"

    c = canvas.Canvas(pdf_file)

    c.setFont("Helvetica-Bold", 18)
    c.drawString(180, 800, "AI Finance Report")

    c.setFont("Helvetica", 12)

    c.drawString(50, 740, f"User : {session['user_name']}")
    c.drawString(50, 710, f"Total Expense : Rs. {total}")
    c.drawString(50, 680, f"Budget : Rs. {budget}")
    c.drawString(50, 650, f"Remaining : Rs. {remaining}")

    c.drawString(50, 600, "Generated by AI Powered Finance Tracker")

    c.save()

    return send_file(pdf_file, as_attachment=True)

@app.route("/export_excel")
def export_excel():

    if "user_id" not in session:
        return redirect("/login")

    mycursor = conn.cursor()

    sql = """
    SELECT expense_date, category, description, amount
    FROM expenses
    WHERE user_id=?
    ORDER BY expense_date DESC
    """

    mycursor.execute(sql, (session["user_id"],))

    data = mycursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    ws.append(["Date", "Category", "Description", "Amount"])

    for row in data:
        ws.append(list(row))
    
    from openpyxl.styles import Font

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["C1"].font = Font(bold=True)
    ws["D1"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 15

    os.makedirs("reports", exist_ok=True)

    filepath = os.path.join("reports", "expense_report.xlsx")
    filepath = "reports/expense_report.xlsx"

    wb.save(filepath)

    mycursor.close()

    return send_file(filepath, as_attachment=True)


# ---------------- RUN APP ----------------

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)